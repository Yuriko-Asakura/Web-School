from django.shortcuts import render

from django.http import JsonResponse
import psutil
from influxdb_client import InfluxDBClient, Point
from influxdb_client.client.write_api import SYNCHRONOUS
import os
from django.contrib.sessions.models import Session

def influxdb_metrics_view(request):
    """Собирает все метрики (CPU, активные сессии, 404 ошибки и т.д.) и записывает их в InfluxDB."""
    # Настройки InfluxDB
    token = os.environ.get("INFLUXDB_TOKEN")
    org = "MPT"
    url = "http://localhost:8086"
    bucket = "met1"

    # Создаём клиент InfluxDB
    client = InfluxDBClient(url=url, token=token, org=org)
    write_api = client.write_api(write_options=SYNCHRONOUS)

    # Собираем метрики CPU
    cpu_percent = psutil.cpu_percent(interval=1)  # Загрузка CPU в процентах
    cpu_count = psutil.cpu_count(logical=True)  # Количество логических ядер CPU
    load_avg = psutil.getloadavg()  # Средняя загрузка системы (1, 5, 15 минут)
    process = psutil.Process()
    process_cpu_usage = process.cpu_percent(interval=0.1)  # Загрузка CPU для текущего процесса

    # Собираем метрики активных сессий
    active_sessions_count = Session.objects.count()

    # Собираем метрики 404 ошибок (если есть)
    # Для этого можно использовать глобальную переменную или кэш, если middleware уже собирает данные
    page_not_found_count = 0  # Здесь можно добавить логику для подсчёта 404 ошибок

    # Создаём точку данных для CPU
    cpu_point = (
        Point("system_metrics")
        .tag("host", "localhost")  # Тег для идентификации хоста
        .field("cpu_percent", cpu_percent)
        .field("cpu_count", cpu_count)
        .field("load_avg_1min", load_avg[0])
        .field("load_avg_5min", load_avg[1])
        .field("load_avg_15min", load_avg[2])
        .field("process_cpu_usage", process_cpu_usage)
    )

    # Создаём точку данных для активных сессий
    sessions_point = (
        Point("session_metrics")
        .tag("host", "localhost")
        .field("active_sessions", active_sessions_count)
    )

    # Создаём точку данных для 404 ошибок
    page_not_found_point = (
        Point("error_metrics")
        .tag("host", "localhost")
        .field("page_not_found", page_not_found_count)
    )

    # Записываем данные в InfluxDB
    write_api.write(bucket=bucket, org=org, record=[cpu_point, sessions_point, page_not_found_point])

    # Возвращаем JSON-ответ
    return JsonResponse({
        "status": "success",
        "message": "Data written to InfluxDB",
        "data": {
            "cpu_metrics": {
                "cpu_percent": cpu_percent,
                "cpu_count": cpu_count,
                "load_avg_1min": load_avg[0],
                "load_avg_5min": load_avg[1],
                "load_avg_15min": load_avg[2],
                "process_cpu_usage": process_cpu_usage,
            },
            "session_metrics": {
                "active_sessions": active_sessions_count,
            },
            "error_metrics": {
                "page_not_found": page_not_found_count,
            },
        }
    })


from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.template import Library
from django.urls import reverse
from django.http import HttpResponseNotFound, HttpResponseRedirect
from django.http import JsonResponse
import bcrypt
import psycopg2
import os
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt 
from django.db import connection
import os
import uuid 
import tempfile
import openpyxl  # Импортируем openpyxl для xlsx
import json  # Импортируем json для json
from datetime import datetime  # Добавляем импорт datetime
import csv  # Импортируем csv для работы с CSV
from django.db import connection, transaction
from django.shortcuts import render, HttpResponse
from django.db import connection, transaction 
import psycopg2 
import random

from prometheus_client import Counter, Histogram, generate_latest
from django.utils import timezone
from django.contrib.auth import login
from django.shortcuts import render, redirect
from .models import Mod_Users

# school_web/views.py
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.conf import settings  # Импортируем settings
from school_web.influxdb_client import get_influxdb_client

def write_data(request):
    client = get_influxdb_client()
    write_api = client.write_api()

    data = {
        "measurement": "temperature",
        "tags": {"location": "room1"},
        "fields": {"value": 22.5},
    }

    # Используем settings.INFLUXDB_SETTINGS
    write_api.write(bucket=settings.INFLUXDB_SETTINGS['met1'], record=data)
    return JsonResponse({"status": "Data written to InfluxDB"})

import psutil
from datetime import datetime
from django.conf import settings
from school_web.influxdb_client import get_influxdb_client
from influxdb_client import InfluxDBClient
from influxdb_client import InfluxDBClient

from django.http import JsonResponse

# --------------------------------------------------------------influxdb
from django.http import JsonResponse
from influxdb_client import InfluxDBClient, Point
from influxdb_client.client.write_api import SYNCHRONOUS
import os
import psutil
from django.contrib.sessions.models import Session
 
from django.http import JsonResponse
from influxdb_client.client.write_api import SYNCHRONOUS
import os
import psutil
import time
from influxdb_client import InfluxDBClient, Point

def influxdb_metrics_view(request):
    """Собирает метрики CPU и записывает их в InfluxDB."""
    token = os.environ.get("INFLUXDB_TOKEN")
    org = "MPT"
    url = "http://localhost:8086"
    bucket = "inf"

    client = InfluxDBClient(url=url, token=token, org=org)
    write_api = client.write_api(write_options=SYNCHRONOUS)

    cpu_percent = psutil.cpu_percent(interval=1)  
    cpu_count = psutil.cpu_count(logical=True)  
    load_avg = psutil.getloadavg()  

    point = (
        Point("cpu_metrics")
        .tag("host", "localhost")  
        .field("cpu_percent", cpu_percent)
        .field("cpu_count", cpu_count)
        .field("load_avg_1min", load_avg[0])
        .field("load_avg_5min", load_avg[1])
        .field("load_avg_15min", load_avg[2])
    )

    write_api.write(bucket=bucket, org=org, record=point)

INFLUX_CONFIG = {
    "url": "http://localhost:8086",
    "token": "H6LB3z6NA7exveKLEMJ5BVBGiJxKDOe6z5bgFNgaK1-nCGHbAvlM8cswE9hbDAHU75Xf1qwbrzmtNZQtK6hxpA==",
    "org": "MPT",
    "bucket": "auth_metrics"
}

def _write_auth_metric(user_id, role, duration=None, login_method=None, client_ip=None, success=None):
    try:
        client = InfluxDBClient(
            url=INFLUX_CONFIG['url'],
            token=INFLUX_CONFIG['token'],
            org=INFLUX_CONFIG['org']
        )
        
        write_api = client.write_api(write_options=SYNCHRONOUS)
        
        point = Point("auth_events") \
            .tag("user_id", str(user_id)) \
            .tag("role", role) \
            .field("count", 1)
        
        if duration is not None:
            point.field("duration", float(duration))
        if login_method is not None:
            point.tag("method", login_method)
        if client_ip is not None:
            point.tag("client_ip", client_ip)
        if success is not None:
            point.tag("success", str(success))
        
        point.time(timezone.now().isoformat())
        
        write_api.write(
            bucket=INFLUX_CONFIG['bucket'],
            record=point
        )
        
        client.close()
    except Exception as e:
        print(f"InfluxDB error: {str(e)}")
                       
def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        try:
            user = Mod_Users.objects.get(Email=email)
        except Mod_Users.DoesNotExist:
            return render(request, 'registration/login.html', {'error': 'Пользователь не найден'})
        if user.check_password(password):
            user.last_login = timezone.now()
            user.save()
            request.session['role'] = user.ID_Role.name
            request.session['user_id'] = user.ID_User
            request.session['email'] = user.Email
            login(request, user)
            if user.ID_Role.name == 'Ученик':
                return redirect('/student/')
            elif user.ID_Role.name == 'Учитель':
                return redirect('/teacher/')
            elif user.ID_Role.name == 'Главный':
                return redirect('/ad/')
            elif user.ID_Role.name == 'Менеджер':
                return redirect('/manager/')
            else:
                return redirect('/teacher/')
        else:
            return render(request, 'registration/login.html', {'error': 'Неверный пароль'})
    else:
        return render(request, 'registration/login.html')
    


   
def main_page(request):
    
    return render(request, 'main_p.html')

from django.shortcuts import render, redirect
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password
Mod_Users = get_user_model()
from django.shortcuts import render, redirect
from .models import Roles, Mod_Users
from django.contrib.auth.hashers import make_password
from django.db import IntegrityError
from .models import Mod_Users, Roles 
from django.contrib import messages
from django.shortcuts import render, redirect

from django.contrib.auth.hashers import make_password

from django.db import IntegrityError
from .models import Mod_Users, Roles  
from django.contrib import messages


def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        role_id = request.POST.get('role_id')
        print(f"username: {username}, email: {email}, password: {password}, role_id: {role_id}")  
        if not username or not email or not password or not role_id:
            messages.error(request, "Все поля обязательны для заполнения.")
            roles = Roles.objects.all()
            return render(request, 'registration/register.html', {'roles': roles})
        if Mod_Users.objects.filter(Username=username).exists() or Mod_Users.objects.filter(Email=email).exists():
            messages.error(request, "Логин или почта уже заняты.")
            roles = Roles.objects.all()
            return render(request, 'registration/register.html', {'roles': roles})
        try:
            role = Roles.objects.get(ID_Role=role_id)
        except Roles.DoesNotExist:
            messages.error(request, "Выбранная роль не существует.")
            roles = Roles.objects.all()
            return render(request, 'registration/register.html', {'roles': roles})
        try:
            hashed_password = make_password(password)  
            user = Mod_Users(
                Username=username,Email=email,Password=hashed_password, ID_Role=role,)
            user.save()
            messages.success(request, "Регистрация прошла успешно! Теперь вы можете войти.")  # Success message
            return redirect('/login/')  # Redirect to the login page
        except IntegrityError as e:
            print(f"IntegrityError: {e}")
            messages.error(request, "Ошибка при сохранении. Возможно, логин или почта уже заняты.")
            roles = Roles.objects.all()
            return render(request, 'registration/register.html', {'roles': roles})

        except Exception as e:
            print(f"Other error: {e}")
            messages.error(request, f"Произошла ошибка: {e}")
            roles = Roles.objects.all()
            return render(request, 'registration/register.html', {'roles': roles})

    else:
        roles = Roles.objects.all()
        print(f"Roles: {list(roles)}")  # Debug output
        return render(request, 'registration/register.html', {'roles': roles})
    
    
    
from django.contrib.auth import logout

def logout_view(request):
    # Удаление данных из сессии
    if 'role' in request.session:
        del request.session['role']
    if 'user_id' in request.session:
        del request.session['user_id']
    if 'email' in request.session:
        del request.session['email']

    # Выход из системы
    logout(request)
    return redirect('/login/')

from django.contrib.auth import authenticate, login
from .models import Mod_Users
 
        
def student_view(request):
    if request.session.get('role') == 'Ученик':
        email = request.session.get('email')
        if email:
            try:
                user = Mod_Users.objects.get(Email=email)
                context = {
                    'email': user.Email,
                    'username': user.Username,
                    'logged_in': True
                }
                return render(request, 'student/profile.html', context)
            except Mod_Users.DoesNotExist:
                return render(request, 'student/profile.html', {'error': 'Пользователь не найден'})
            except Exception as e:
                return render(request, 'student/profile.html', {'error': f'Ошибка базы данных: {str(e)}'})
        else:
            return redirect('/login/')
    else:
        return redirect('/login/')
    
def teacher_view(request):
    if 'role' in request.session and request.session['role'] == 'Учитель':
        email = request.session.get('email')
        if email:
            try:
                user = Mod_Users.objects.get(Email=email) #Использование ORM
                context = {
                    'email': user.Email,
                    'username': user.Username,
                    'first_name': user.ID_User,  # Предполагаю, что это ID
                    'last_name': user.ID_User,  # Добавьте заглушку
                    'logged_in': True
                }
                return render(request, 'teacher/profile.html', context)
            except Mod_Users.DoesNotExist:
                return HttpResponse("User not found.")
            except Exception as e:
                return HttpResponse(f"An error occurred: {e}")
        else:
            return redirect('/login/')
    else:
        return redirect('/login/')

def admin_view(request):
    if request.session.get('role') == 'Главный':
        email = request.session.get('email')
        if email:
            try:
                user = Mod_Users.objects.get(Email=email)
                context = {
                    'email': user.Email,
                    'username': user.Username,
                    'first_name': user.ID_User,
                    'last_name': user.Username, # Replace with the actual field if different
                    'logged_in': True
                }
                return render(request, 'ad/profile.html', context)
            except Mod_Users.DoesNotExist:
                return render(request, 'ad/profile.html', {'error': 'Пользователь не найден'}) 
            except Exception as e:
                return render(request, 'ad/profile.html', {'error': f'Ошибка базы данных: {str(e)}'}) 
        else:
            return render(request, '/login/', {'error': 'Ошибка: email не найден в сессии.'}) 
    else:
        return redirect('/login/')
    
    
def manager_view(request):
    if 'role' in request.session and request.session['role'] == 'Менеджер':
        # Получение информации о пользователе
        email = request.session.get('email')  # Получение email из сессии
        if email:
            # Получение информации о пользователе из базы данных
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT * FROM Users WHERE Email = %s",
                    [email]
                )
                user = cursor.fetchone()

            if user:
                context = {
                    'email': user[2],  # Индекс 2 - это Email в таблице Users
                    'username': user[1],  # Индекс 1 - это Username в таблице Users
                    'first_name': user[0],  # Индекс 0 - это ID_User в таблице Users
                    'last_name': user[0],  #  Добавьте заглушку, пока вы не определите столбец для имени
                    'logged_in': True
                }
                return render(request, 'manager/profile.html', context)
        else:
            return redirect('/login/')  
    else:
        return redirect('/login/')

from django.shortcuts import render, redirect
from .models import Courses, Predmet
def courses_view(request):
    user_id = request.session.get('user_id')  # Получаем ID пользователя из сессии
    if not user_id:
        return redirect('/login/')  # Если пользователь не авторизован, перенаправляем на страницу входа

    try:
        # Получаем объект пользователя
        user = Mod_Users.objects.get(ID_User=user_id)
    except Mod_Users.DoesNotExist:
        return redirect('/login/')  # Если пользователь не найден, перенаправляем на страницу входа

    # Получаем все курсы пользователя с использованием ORM и предварительно загружаем связанные объекты Predmet
    courses = Courses.objects.filter(ID_User=user).select_related('Predmet')

    # Преобразуем данные в удобный формат
    courses_data = []
    for course in courses:
        course_data = {
            'ID_Course': course.ID_Course,
            'PredmetName': course.Predmet.PredmetName if course.Predmet else None,
            'Description': course.Description,
            'CourseName': course.CourseName,
            'ID_User': course.ID_User.ID_User,  # Используем ID_User из объекта Mod_Users
            'IsBlocked': course.IsBlocked,
            'IsSubscribed': course.IsSubscribed,
        }
        courses_data.append(course_data)

    # Получаем все предметы для отображения
    all_subjects = Predmet.objects.all()
    subjects_dict = {subject.ID_Predmet: subject.PredmetName for subject in all_subjects}

    # Передаем данные в шаблон
    context = {
        'courses': courses_data,
        'subjects_dict': subjects_dict,
    }
    return render(request, 'courses/courses.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Courses, Predmet, Mod_Users, Topics, Tests, Complexity

# Курсы
def create_course_view(request):
    predmets = Predmet.objects.all()
    
    if request.method == 'POST':
        course_name = request.POST.get('course_name')
        predmet_id = request.POST.get('predmet')
        description = request.POST.get('description')
        user_id = request.session.get('user_id')

        if not all([course_name, predmet_id, description, user_id]):
            messages.error(request, "Все поля обязательны для заполнения")
            return render(request, 'courses/create_course.html', {'predmets': predmets})

        if Courses.objects.filter(CourseName=course_name).exists():
            messages.error(request, "Курс с таким названием уже существует")
            return render(request, 'courses/create_course.html', {'predmets': predmets})

        try:
            user = Mod_Users.objects.get(ID_User=user_id)
            predmet = Predmet.objects.get(ID_Predmet=predmet_id)
            
            course = Courses.objects.create(
                CourseName=course_name,
                Predmet=predmet,
                Description=description,
                ID_User=user
            )
            messages.success(request, "Курс успешно создан")
            # Вместо редиректа снова показываем форму (но очищенную)
            return render(request, 'courses/create_course.html', {'predmets': predmets})
            
        except (Mod_Users.DoesNotExist, Predmet.DoesNotExist) as e:
            messages.error(request, f"Ошибка: {str(e)}")
            return render(request, 'courses/create_course.html', {'predmets': predmets})

    return render(request, 'courses/create_course.html', {'predmets': predmets})
def update_course_view(request, course_id):
    course = get_object_or_404(Courses, ID_Course=course_id)
    predmets = Predmet.objects.all()

    if request.method == 'POST':
        course.CourseName = request.POST.get('course_name')
        course.Predmet_id = request.POST.get('predmet')
        course.Description = request.POST.get('description')
        course.save()
        messages.success(request, "Курс успешно обновлен")
        return redirect('courses_list')

    return render(request, 'courses/update_course.html', {
        'course': course,
        'predmets': predmets
    })

def delete_course_view(request, course_id):
    course = get_object_or_404(Courses, ID_Course=course_id)
    course.delete()
    messages.success(request, "Курс успешно удален")
    return redirect('courses_list')

# Темы
def course_topics_view(request, course_id):
    course = get_object_or_404(Courses, ID_Course=course_id)
    topics = course.topics_set.all()  # Используем related_name
    return render(request, 'courses/topics.html', {
        'topics': topics,
        'course': course
    })

def create_topic_view(request, course_id):
    course = get_object_or_404(Courses, ID_Course=course_id)
    
    if request.method == 'POST':
        topic_name = request.POST.get('topic_name')
        if not topic_name:
            messages.error(request, "Название темы обязательно")
            return redirect('create_topic', course_id=course_id)
            
        Topics.objects.create(TopicName=topic_name, ID_Course=course)
        messages.success(request, "Тема успешно создана")
        return redirect('course_topics', course_id=course_id)
        
    return render(request, 'courses/create_topic.html', {'course': course})

def update_topic_view(request, course_id, topic_id):
    topic = get_object_or_404(Topics, ID_Topic=topic_id, ID_Course_id=course_id)
    
    if request.method == 'POST':
        topic.TopicName = request.POST.get('topic_name')
        topic.save()
        messages.success(request, "Тема успешно обновлена")
        return redirect('course_topics', course_id=course_id)
        
    return render(request, 'courses/update_topic.html', {
        'topic': topic,
        'course_id': course_id
    })

def delete_topic_view(request, course_id, topic_id):
    topic = get_object_or_404(Topics, ID_Topic=topic_id, ID_Course_id=course_id)
    topic.delete()
    messages.success(request, "Тема успешно удалена")
    return redirect('course_topics', course_id=course_id)

# Тесты
def tests_for_topic_view(request, course_id, topic_id):
    topic = get_object_or_404(Topics, ID_Topic=topic_id, ID_Course_id=course_id)
    tests = topic.tests_set.all()  # Используем related_name
    return render(request, 'courses/tests/tests.html', {
        'tests': tests,
        'topic': topic,
        'course_id': course_id
    })

def create_test_view(request, course_id, topic_id):
    topic = get_object_or_404(Topics, ID_Topic=topic_id, ID_Course_id=course_id)
    complexities = Complexity.objects.all()
    
    if request.method == 'POST':
        test_name = request.POST.get('test_name')
        description = request.POST.get('description')
        complexity_id = request.POST.get('complexity')
        
        if not all([test_name, description, complexity_id]):
            messages.error(request, "Все поля обязательны для заполнения")
            return redirect('create_test', course_id=course_id, topic_id=topic_id)
            
        try:
            complexity = Complexity.objects.get(ID_Complexity=complexity_id)
            Tests.objects.create(
                TestName=test_name,
                Description=description,
                ID_Complexity=complexity,
                ID_Topic=topic
            )
            messages.success(request, "Тест успешно создан")
            return redirect('tests_for_topic', course_id=course_id, topic_id=topic_id)
            
        except Complexity.DoesNotExist:
            messages.error(request, "Выбранная сложность не существует")
            return redirect('create_test', course_id=course_id, topic_id=topic_id)
    
    return render(request, 'courses/tests/create_test.html', {
        'topic': topic,
        'complexities': complexities,
        'course_id': course_id, #or just course_id if correct
    })

def update_test_view(request, course_id, test_id, topic_id):
    test = get_object_or_404(Tests, ID_Test=test_id, ID_Topic_id=topic_id)
    complexities = Complexity.objects.all()
    
    if request.method == 'POST':
        test.TestName = request.POST.get('test_name')
        test.Description = request.POST.get('description')
        test.ID_Complexity_id = request.POST.get('complexity')
        test.save()
        messages.success(request, "Тест успешно обновлен")
        return redirect('tests_for_topic', course_id=course_id, topic_id=topic_id)
        
    return render(request, 'courses/tests/update_test.html', {
        'test': test,
        'complexities': complexities,
        'topic_id': topic_id
    })

def delete_test_view(request, course_id, test_id, topic_id):
    test = get_object_or_404(Tests, ID_Test=test_id, ID_Topic_id=topic_id)
    test.delete()
    messages.success(request, "Тест успешно удален")
    return redirect('tests_for_topic', course_id=course_id, topic_id=topic_id)

#Вопросы добавление


from django.shortcuts import render, redirect
from .models import Questions, Tests, QuestionTypes

from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Tests, Questions, QuestionTypes

from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Tests, Questions, QuestionTypes

def create_question_view(request, course_id, topic_id, test_id):
    if request.method == 'POST':
        question_text = request.POST.get('question_text')
        correct_answer = request.POST.get('correct_answer')
        question_type = request.POST.get('question_type')
        num_options = int(request.POST.get('num_options'))

        # Проверка типа вопроса и количества вариантов
        if question_type == '1' and num_options < 2:
            return render(request, 'courses/tests/create_question_multiple.html', {
                'course_id': course_id,
                'topic_id': topic_id,
                'test_id': test_id,
                'error_message': 'Для типа вопроса "Выбор одного ответа" необходимо минимум 2 варианта ответа.',
            })

        # Получение списка правильных ответов
        correct_answers = []
        for i in range(num_options):
            if request.POST.get(f'correct_answer_{i}'):
                correct_answers.append(request.POST.get(f'wrong_answer_{i}'))  # Предполагаем, что правильный ответ также хранится в wrong_answer

        # Соединяем правильные ответы с разделителем
        correct_answer_str = '♡'.join(correct_answers)

        # Получение списка неправильных ответов
        wrong_answers = []
        for i in range(num_options):
            if not request.POST.get(f'correct_answer_{i}'):
                wrong_answers.append(request.POST.get(f'wrong_answer_{i}'))

        # Соединяем неправильные ответы с разделителем
        wrong_answer_str = '♡'.join(wrong_answers)

        # Получение объекта теста и типа вопроса
        test = Tests.objects.get(ID_Test=test_id)
        question_type_obj = QuestionTypes.objects.get(ID_QuestionType=question_type)

        # Создание и сохранение вопроса с использованием ORM
        question = Questions(
            ID_Test=test,
            QuestionText=question_text,
            correct_answer=correct_answer_str,
            wrong_answer=wrong_answer_str,
            ID_QuestionType=question_type_obj,
        )
        question.save()

        # Отправляем успешный ответ с URL для перенаправления
        return JsonResponse({'redirect_url': f'/courses/{course_id}/topics/{topic_id}/tests/{test_id}/questions/'})
    else:
        question_type = request.GET.get('question_type')
        if question_type == '1':
            return redirect(f'/courses/{course_id}/topics/{topic_id}/tests/{test_id}/questions/create_multiple/')
        elif question_type == '2':
            return redirect(f'/courses/{course_id}/topics/{topic_id}/tests/{test_id}/questions/create_text/')
        else:
            return render(request, 'courses/tests/create_question.html', {
                'course_id': course_id,
                'topic_id': topic_id,
                'test_id': test_id,
            })

from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Questions, Tests, QuestionTypes

def create_question_multiple_view(request, course_id, topic_id, test_id):
    test = Tests.objects.get(ID_Test=test_id)
    if request.method == 'POST':
        question_text = request.POST.get('question_text')
        question_type_id = request.POST.get('question_type')
        num_options = int(request.POST.get('num_options'))

        # Проверка типа вопроса и количества вариантов
        if question_type_id == '1' and num_options < 2:
            return render(request, 'courses/tests/create_question_multiple.html', {
                'course_id': course_id,
                'topic_id': topic_id,
                'test_id': test_id,
                'error_message': 'Для типа вопроса "Выбор одного ответа" необходимо минимум 2 варианта ответа.',
            })

        # Получение списка правильных ответов
        correct_answers = []
        for i in range(num_options):
            if request.POST.get(f'correct_answer_{i}'):
                correct_answers.append(request.POST.get(f'wrong_answer_{i}'))  # Предполагаем, что правильный ответ также хранится в wrong_answer

        # Соединяем правильные ответы с разделителем
        correct_answer_str = '♡'.join(correct_answers)

        # Получение списка неправильных ответов
        wrong_answers = []
        for i in range(num_options):
            if not request.POST.get(f'correct_answer_{i}'):
                wrong_answers.append(request.POST.get(f'wrong_answer_{i}'))

        # Соединяем неправильные ответы с разделителем
        wrong_answer_str = '♡'.join(wrong_answers)

        # Получаем объект типа вопроса
        question_type = QuestionTypes.objects.get(ID_QuestionType=question_type_id)

        # Создаем вопрос
        question = Questions(
            QuestionText=question_text,
            correct_answer=correct_answer_str,
            wrong_answer=wrong_answer_str,
            ID_Test=test,
            ID_QuestionType=question_type,
        )
        question.save()

        # Отправляем успешный ответ с URL для перенаправления
        return redirect('questions_for_test', course_id=course_id, topic_id=topic_id, test_id=test.ID_Test)

    else:
        return render(request, 'courses/tests/create_question_multiple.html', {
            'course_id': course_id,
            'topic_id': topic_id,
            'test_id': test_id,
        })

from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Questions, Tests, QuestionTypes

def create_question_text_view(request, course_id, topic_id, test_id):
    test = Tests.objects.get(ID_Test=test_id)
    if request.method == 'POST':
        question_text = request.POST.get('question_text')
        correct_answer = request.POST.get('correct_answer')
        question_type_id = request.POST.get('question_type')

        # Получаем объект типа вопроса
        question_type = QuestionTypes.objects.get(ID_QuestionType=question_type_id)

        # Создаем вопрос
        question = Questions(
            QuestionText=question_text,
            correct_answer=correct_answer,
            wrong_answer='',
            ID_Test=test,
            ID_QuestionType=question_type,
        )
        question.save()

        # Отправляем успешный ответ с URL для перенаправления
        return redirect('questions_for_test', course_id=course_id, topic_id=topic_id, test_id=test.ID_Test)

    else:
        return render(request, 'courses/tests/create_question_text.html', {
            'course_id': course_id,
            'topic_id': topic_id,
            'test_id': test_id,
        })
      
from django.shortcuts import render, redirect, get_object_or_404
from .models import Questions, QuestionTypes

def update_question_view(request, course_id, topic_id, test_id, question_id):
    question = get_object_or_404(Questions, ID_Question=question_id)
    question_types = QuestionTypes.objects.all()
    if request.method == 'POST':
        question_text = request.POST.get('question_text')
        correct_answer = request.POST.get('correct_answer')
        question_type_id = request.POST.get('question_type')
        question_type = QuestionTypes.objects.get(ID_QuestionType=question_type_id)
        question.QuestionText = question_text
        question.correct_answer = correct_answer
        question.ID_QuestionType = question_type
        question.save()
        return redirect(f'/courses/{course_id}/topics/{topic_id}/tests/{test_id}/questions/')
    else:
        context = {
            'question': question,
            'course_id': course_id,
            'topic_id': topic_id,
            'test_id': test_id,
            'question_types': question_types,
        }
        return render(request, 'courses/tests/update_question.html', context)   
    
from django.shortcuts import redirect, get_object_or_404
from .models import Questions

def delete_question_view(request, course_id, topic_id, test_id, question_id):
    question = get_object_or_404(Questions, ID_Question=question_id)
    question.delete()
    return redirect(f'/courses/{course_id}/topics/{topic_id}/tests/{test_id}/questions/')

from django.shortcuts import render, get_object_or_404
from .models import Questions, Tests

def questions_for_test_view(request, course_id, topic_id, test_id):
    test = get_object_or_404(Tests, ID_Test=test_id)
    questions = Questions.objects.filter(ID_Test=test)
    return render(request, 'courses/tests/questions.html', {
        'questions': questions,
        'course_id': course_id,
        'topic_id': topic_id,
        'test_id': test_id,
    })
    
#Добавление для лекций

# myapp/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import Http404
from django.db import connection
from django.contrib import messages
from django.utils import timezone
# Helper function to fetch Lecture data using raw SQL
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from django.db import connection
from django.http import Http404
from django.utils import timezone

class Lecture:
    def __init__(self, id, title, content):
        self.id = id
        self.title = title
        self.content = content

def fetch_lecture_by_id(cursor, lecture_id):
    cursor.execute("SELECT ID_Lecture, Title, Content FROM Lectures WHERE ID_Lecture = %s", [lecture_id])
    lecture_data = cursor.fetchone()
    if lecture_data:
        return Lecture(id=lecture_data[0], title=lecture_data[1], content=lecture_data[2])
    return None


from django.shortcuts import render
from .models import Lectures, Topics, Courses

def lecture_list(request, course_id, topic_id):
    course = Courses.objects.get(ID_Course=course_id)
    topic = Topics.objects.get(ID_Topic=topic_id)
    lectures = Lectures.objects.filter(ID_Topic=topic)
    return render(request, 'lectures/lecture_list.html', {
        'lectures': lectures,
        'course_id': course_id,
        'topic_id': topic_id,
    })

from django.shortcuts import render, redirect
from .models import Lectures, Topics

import logging
from django.shortcuts import render, redirect
from .models import Topics, Lectures

logger = logging.getLogger(__name__)

def create_lecture(request, course_id, topic_id):
    topic = Topics.objects.get(ID_Topic=topic_id)
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        logger.info(f"Title: {title}, Content: {content}")  # Логирование
        try:
            lecture = Lectures(
                Title=title,
                Content=content,
                ID_Topic=topic,
            )
            lecture.full_clean()  # Валидация модели
            lecture.save()
            logger.info("Lecture saved successfully")  # Логирование
            return redirect(f'/courses/{course_id}/topics/{topic_id}/lectures/')
        except ValidationError as e:
            logger.error(f"Validation error: {e}")  # Логирование ошибок валидации
        except Exception as e:
            logger.error(f"Error saving lecture: {e}")  # Логирование других ошибок
    else:
        return render(request, 'lectures/create_lecture.html', {
            'course_id': course_id,
            'topic_id': topic_id,
        })


from django.shortcuts import render, redirect, get_object_or_404
from .models import Lectures

def edit_lecture(request, course_id, topic_id, pk):
    lecture = get_object_or_404(Lectures, ID_Lecture=pk)
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        lecture.Title = title
        lecture.Content = content
        lecture.save()
        return redirect('lecture_detail', course_id=course_id, topic_id=topic_id, pk=lecture.ID_Lecture)
    else:
        return render(request, 'lectures/edit_lecture.html', {
            'lecture': lecture,
            'course_id': course_id,
            'topic_id': topic_id,
        })

from django.shortcuts import render, get_object_or_404
from .models import Lectures

def lecture_detail(request, course_id, topic_id, pk):
    lecture = get_object_or_404(Lectures, ID_Lecture=pk)
    return render(request, 'lectures/lecture_detail.html', {
        'lecture': lecture,
        'course_id': course_id,
        'topic_id': topic_id,
    })


#Admin действия для него


def create_database_backup(database_name, file_format="xlsx"):
    """Создает резервную копию базы данных PostgreSQL и сохраняет ее в файл.

    Args:
        database_name: Имя базы данных для резервного копирования.
        file_format: Формат файла для резервной копии ('xlsx', 'json', 'sql', 'csv').
    """

    # Получаем путь к папке "загрузки" (предполагаем, что это стандартный путь)
    download_path = os.path.join(os.path.expanduser("~"), "Downloads")
    backup_path = os.path.join(download_path, f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_format}")

    with connection.cursor() as cursor:
        with open(backup_path, "w", encoding="utf-8") as f:
            # Получаем список всех таблиц в базе данных
            cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = 'public'")
            table_names = cursor.fetchall()

            # Создаем файл Excel (xlsx), JSON, SQL или CSV
            if file_format == "xlsx":
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["Таблица", "Столбец", "Значение"])  # Добавляем заголовок
            elif file_format == "json":
                data = {}
                for table_name in table_names:
                    data[table_name[0]] = []
            elif file_format == "sql":
                f.write("-- Резервная копия базы данных: " + database_name + "\n\n")
            elif file_format == "csv":
                writer = csv.writer(f)

            # Делаем резервную копию каждой таблицы
            for table_name in table_names:
                table_name = table_name[0]
                cursor.execute(f"SELECT * FROM {table_name}")
                rows = cursor.fetchall()

                if file_format == "xlsx":
                    # Заполняем файл Excel
                    for row in rows:
                        for i, value in enumerate(row):
                            # Убираем временную зону из datetime
                            if isinstance(value, datetime):
                                value = value.replace(tzinfo=None)
                            sheet.append([table_name, cursor.description[i][0], value])
                elif file_format == "json":
                    # Заполняем файл JSON
                    for row in rows:
                        # Преобразуем datetime в строку для JSON
                        row = [str(x) if isinstance(x, datetime) else x for x in row]
                        data[table_name].append(dict(zip([col[0] for col in cursor.description], row)))
                elif file_format == "sql":
                    # Заполняем файл SQL
                    f.write(f"-- Таблица: {table_name}\n")
                    f.write(f"COPY {table_name} FROM STDIN WITH (FORMAT CSV, HEADER, DELIMITER ',');\n")
                    for row in rows:
                        f.write(",".join([str(x) for x in row]) + "\n")
                    f.write(";\n\n")
                elif file_format == "csv":
                    # Заполняем файл CSV
                    writer.writerow([cursor.description[i][0] for i in range(len(cursor.description))])
                    for row in rows:
                        writer.writerow(row)

            # Сохраняем файл
            if file_format == "xlsx":
                workbook.save(backup_path)
            elif file_format == "json":
                json.dump(data, f, indent=4)
            elif file_format == "sql" or file_format == "csv":
                # Файл уже сохранен в цикле
                pass

        print(f"Резервная копия базы данных {database_name} успешно создана в {backup_path}.")

    with open(backup_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(backup_path)}"'
        return response

@csrf_exempt
def backup_database(request):
    if request.method == 'POST':
        database_name = 'School'
        file_format = request.POST.get('file_format')
        
        if file_format:
            return create_database_backup(database_name, file_format=file_format)
        else:
            return HttpResponse('Не выбран формат файла.')
    else:
        return render(request, 'ad/backup_and_import.html')

DATABASES_SETTINGS = {
    'ENGINE': 'django.db.backends.postgresql',
    'NAME': 'School',
    'USER': 'postgres',
    'PASSWORD': '2005',
    'HOST': 'localhost',
    'PORT': '5432',
}


def import_database_backup(backup_path, file_format):
    """Импортирует резервную копию базы данных PostgreSQL из файла."""

    try:
        conn = psycopg2.connect(
            database=DATABASES_SETTINGS['NAME'],
            user=DATABASES_SETTINGS['USER'],
            password=DATABASES_SETTINGS['PASSWORD'],
            host=DATABASES_SETTINGS['HOST'],
            port=DATABASES_SETTINGS['PORT'],
        )
        cur = conn.cursor()

        if file_format == "xlsx":
            workbook = openpyxl.load_workbook(backup_path, read_only=True, data_only=True) #data_only=True для получения значений, а не формул
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                table_name = sheet_name # Предполагаем, что имя листа - это имя таблицы
                header = [cell.value for cell in sheet[1]] # заголовок в первой строке
                for row in sheet.iter_rows(min_row=2):
                    values = [cell.value for cell in row]
                    placeholders = ", ".join(["%s"] * len(values))
                    cur.execute(f"INSERT INTO {table_name} ({', '.join(header)}) VALUES ({placeholders})", values)

        elif file_format == "json":
            with open(backup_path, 'r') as f:
                data = json.load(f)
                for table_name, rows in data.items():
                    for row in rows:
                        columns = ", ".join(row.keys())
                        placeholders = ", ".join(["%s"] * len(row))
                        values = tuple(row.values())
                        cur.execute(f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})", values)


        elif file_format == "sql":
            with open(backup_path, 'r', encoding='utf-8') as f: # Добавлена кодировка
                cur.execute("BEGIN;") # Begin transaction
                cur.execute(f.read())
                cur.execute("COMMIT;") # Commit transaction


        elif file_format == "csv":
            with open(backup_path, 'r', encoding='utf-8') as f: # Добавлена кодировка
                reader = csv.reader(f)
                header = next(reader)
                table_name = 'your_table_name' #  !!!  Замените 'your_table_name' на фактическое имя вашей таблицы  !!!
                placeholders = ", ".join(["%s"] * len(header))
                for row in reader:
                    cur.execute(f"INSERT INTO {table_name} ({', '.join(header)}) VALUES ({placeholders})", row)

        conn.commit()
        print(f"Данные успешно импортированы из {backup_path}.")
        return True

    except psycopg2.Error as e:
        print(f"Ошибка PostgreSQL: {e}")
        conn.rollback()
        return False
    except (json.JSONDecodeError, openpyxl.utils.exceptions.InvalidFileException, csv.Error) as e:
        print(f"Ошибка обработки файла: {e}")
        return False
    except Exception as e:
        print(f"Общая ошибка: {e}")
        return False
    finally:
        if conn:
            cur.close()
            conn.close()

@csrf_exempt
def import_database(request):
    if request.method == 'POST':
        uploaded_file = request.FILES['backup_file']
        file_format = request.POST.get('file_format')

        if not uploaded_file:
            return HttpResponse("Файл не выбран.")

        temp_dir = tempfile.gettempdir()
        file_name, file_ext = os.path.splitext(uploaded_file.name)
        unique_id = uuid.uuid4()
        file_path = os.path.join(temp_dir, f"{file_name}_{unique_id}{file_ext}")
        with open(file_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        if import_database_backup(file_path, file_format):
            os.remove(file_path)
            return HttpResponse("Данные успешно импортированы.")
        else:
            os.remove(file_path)
            return HttpResponse("Ошибка при импорте данных.")
    else:
        return render(request, 'ad/import.html') 
     
def all_stats_view(request):
    user_id = request.session.get('user_id')
    if user_id:
        try:
            with connection.cursor() as cursor:
                # Courses per subject
                cursor.execute("""
                    SELECT p.PredmetName, COUNT(DISTINCT c.ID_Course) AS course_count
                    FROM Courses c
                    JOIN Predmet p ON c.Predmet = p.ID_Predmet
                    WHERE c.ID_User = %s
                    GROUP BY p.PredmetName
                """, [user_id])
                courses_data = cursor.fetchall()

                # Roles and users count
                cursor.execute("""
                    SELECT r.RoleName, COUNT(*) AS user_count
                    FROM Roles r
                    JOIN Users u ON r.ID_Role = u.ID_Role
                    GROUP BY r.RoleName
                """)
                role_data = cursor.fetchall()

            # Prepare data for Chart.js
            subject_names = [row[0] for row in courses_data]
            subject_counts = [row[1] for row in courses_data]

            role_names = [row[0] for row in role_data]
            role_counts = [row[1] for row in role_data]

            combined_chart_data = {
                'labels': subject_names + role_names,
                'datasets': [
                    {
                        'label': 'Количество курсов',
                        'data': subject_counts + [0] * len(role_names),  # Pad with zeros
                        'backgroundColor': '#009c7b',
                    },
                    {
                        'label': 'Количество пользователей по ролям',
                        'data': [0] * len(subject_names) + role_counts,  # Pad with zeros
                        'backgroundColor': '#36A2EB',
                    },
                ]
            }

            return render(request, 'ad/all_course_stats.html', {
                'combined_chart_data': json.dumps(combined_chart_data),
            })

        except Exception as e:
            print(f"Ошибка базы данных: {e}")
            return render(request, 'error.html', {'error_message': str(e)})

    else:
        return redirect('/login/')


#Менеджер роль

# def update_topic_view(request, course_id, topic_id):
#     if request.method == 'POST':
#         topic_name = request.POST.get('topic_name')
#         user_id = request.user.id  # Получаем ID пользователя

#         with connection.cursor() as cursor:
#             # Проверяем роль пользователя
#             cursor.execute(
#                 "SELECT ID_Role FROM Users WHERE ID_User = %s", [user_id]
#             )
#             role_id = cursor.fetchone()[0]
#             cursor.execute(
#                 "SELECT RoleName FROM Roles WHERE ID_Role = %s", [role_id]
#             )
#             role_name = cursor.fetchone()[0]

#             if role_name == 'Менеджер':
#                 # Обновление данных только если пользователь - менеджер
#                 cursor.execute(
#                     "UPDATE Topics SET TopicName = %s WHERE ID_Topic = %s",
#                     [topic_name, topic_id]
#                 )

#                 # Записываем аудит 
#                 cursor.execute(
#                     "INSERT INTO AuditLog (TableName, Operation, UserID, Data, ID_Course, ID_Topic, ID_Test) VALUES (%s, %s, %s, %s, %s, %s, %s)",
#                     ['Topics', 'UPDATE', user_id, json_build_object('old_data', json_build_object('TopicName', request.POST.get('old_topic_name')), 'new_data', json_build_object('TopicName', topic_name)), course_id, topic_id, None]  # None - если изменение темы не связано с тестом
#                 )
#                 connection.commit()
#             else:
#                 return HttpResponse("У вас нет прав на изменение темы.")

#         return redirect(f'/courses/{course_id}/topics/{topic_id}/')
#     else:
#         with connection.cursor() as cursor:
#             cursor.execute(
#                 "SELECT TopicName FROM Topics WHERE ID_Topic = %s", [topic_id]
#             )
#             topic_name = cursor.fetchone()[0]
#         return render(request, 'courses/topics/update_topic.html', {
#             'course_id': course_id,
#             'topic_id': topic_id,
#             'topic_name': topic_name,
#         })


#Представления для ученика 


from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db.models import Q

from django.shortcuts import render
from django.db.models import Exists, OuterRef
from .models import Courses, Predmet, UserCourses

def courses_list_view(request):
    user_id = request.session.get('user_id')
    
    # Получаем все курсы с аннотацией статуса подписки (используем другое имя для аннотации)
    courses = Courses.objects.annotate(
        user_is_subscribed=Exists(  # Изменили имя аннотации
            UserCourses.objects.filter(
                ID_User=user_id,
                ID_Course=OuterRef('ID_Course')
            )
        )
    ).select_related('Predmet').order_by('CourseName')
    
    # Подготавливаем данные для шаблона
    courses_data = []
    for course in courses:
        courses_data.append({
            'ID_Course': course.ID_Course,
            'CourseName': course.CourseName,
            'Description': course.Description,
            'PredmetName': course.Predmet.PredmetName if course.Predmet else '',
            # Используем аннотацию user_is_subscribed вместо поля модели IsSubscribed
            'IsSubscribed': course.user_is_subscribed,
            'ImageUrl': course.ImageUrl if hasattr(course, 'ImageUrl') else ''
        })

    return render(request, 'student/courses/courses_list.html', {
        'courses': courses_data,
    })
    
from django.shortcuts import redirect, render
from django.db import transaction
from .models import Courses, UserCourses, Predmet, Topics

def subscribe_course_view(request, course_id):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')  # Перенаправляем если пользователь не авторизован

    try:
        with transaction.atomic():
            # Проверяем и обновляем подписку в одной транзакции
            subscription, created = UserCourses.objects.get_or_create(
                ID_User_id=user_id,
                ID_Course_id=course_id
            )
            
            if not created:
                subscription.delete()
                message = "Вы успешно отписались от курса."
            else:
                message = "Вы успешно подписались на курс."
                
    except Exception as e:
        message = f"Произошла ошибка: {str(e)}"

    return redirect('courses_list')

def subscribed_courses_view(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('login')

    # Получаем подписанные курсы с аннотацией тем
    subscribed_courses = Courses.objects.filter(
        usercourses__ID_User=user_id
    ).select_related('Predmet').prefetch_related('topics_set').order_by('CourseName')

    # Формируем данные для шаблона
    courses = []
    for course in subscribed_courses:
        courses.append({
            'ID_Course': course.ID_Course,
            'CourseName': course.CourseName,
            'Description': course.Description,
            'PredmetName': course.Predmet.PredmetName if course.Predmet else '',
            'topics': list(course.topics_set.values_list('TopicName', flat=True)),
        })

    return render(request, 'student/courses/subscribed_courses.html', {
        'courses': courses,
    })
from django.shortcuts import render, get_object_or_404
from .models import Courses, Topics, Predmet

def course_details_view(request, course_id):
    # Получаем курс или возвращаем 404
    course = get_object_or_404(Courses, ID_Course=course_id)
    
    # Получаем все темы для этого курса
    topics = Topics.objects.filter(ID_Course=course_id).order_by('ID_Topic')
    
    context = {
        'course': {
            'ID_Course': course.ID_Course,
            'CourseName': course.CourseName,
            'Description': course.Description,
            'PredmetName': course.Predmet.PredmetName if course.Predmet else '',
            'topics': [{'ID_Topic': t.ID_Topic, 'TopicName': t.TopicName} for t in topics]
        }
    }
    return render(request, 'student/courses/course_details.html', context)

from django.shortcuts import render, HttpResponse
from .models import Topics, Tests

def topic_tests_view(request, topic_id):
    try:
        topic = Topics.objects.get(ID_Topic=topic_id)
        tests = Tests.objects.filter(ID_Topic=topic_id)
        
        if not tests.exists():
            return HttpResponse('Тесты для этой темы не найдены.', status=404)
            
        return render(request, 'student/courses/topic_tests.html', {
            'topic_name': topic.TopicName,
            'tests': [{
                'ID_Test': test.ID_Test,
                'TestName': test.TestName,
                'Description': test.Description
            } for test in tests],
            'topic_id': topic_id
        })
    except Topics.DoesNotExist:
        return HttpResponse('Тема не найдена.', status=404)      

from django.http import HttpResponseNotFound      
from django.shortcuts import render, redirect
from django.db import transaction
from .models import Tests, Questions, Topics, Courses, UserTests
from datetime import datetime
import random

def start_test_view(request, test_id, topic_id):
    try:
        test = Tests.objects.get(ID_Test=test_id)
        questions = Questions.objects.filter(ID_Test=test_id)
        
        if not questions.exists():
            return HttpResponseNotFound("Test not found")

        questions_data = []
        for question in questions:
            correct_answers = [x.strip() for x in question.correct_answer.split('♡') if x.strip()]
            wrong_answers = [x.strip() for x in question.wrong_answer.split('♡') if x.strip()]
            all_answers = correct_answers + wrong_answers
            random.shuffle(all_answers)

            questions_data.append({
                'id': question.ID_Question,
                'text': question.QuestionText,
                'answers': all_answers,
                'type': question.ID_QuestionType
            })

        return render(request, 'student/courses/test_start.html', {
            'test_name': test.TestName,
            'questions': questions_data,
            'test_id': test_id,
            'topic_id': topic_id
        })
    except Tests.DoesNotExist:
        return HttpResponseNotFound("Test not found")

def test_submit(request, topic_id, test_id):
    if request.method == 'POST':
        try:
            answers = {}
            for key, value in request.POST.items():
                if key.startswith('answer_'):
                    parts = key.split('_')
                    question_id = parts[1]
                    answer = value
                    answers.setdefault(question_id, []).append(answer)
                elif key.startswith('text_answer_'):
                    question_id = key.replace('text_answer_', '')
                    answers[question_id] = value

            # Получение ID_Course через ID_Topic
            try:
                topic = Topics.objects.get(ID_Topic=topic_id)
                course_id = topic.ID_Course.ID_Course
            except Topics.DoesNotExist:
                return render(request, 'student/courses/test_results.html', 
                           {'error_message': 'Тема не найдена.'})

            # Получаем вопросы и правильные ответы
            questions = Questions.objects.filter(ID_Test=test_id)
            if not questions.exists():
                return render(request, 'student/courses/test_results.html', 
                           {'error_message': 'Вопросы для этого теста не найдены.'})

            correct_count = 0
            total_questions = questions.count()
            
            for question in questions:
                user_answers = answers.get(str(question.ID_Question), [])
                correct_answers = [x.strip() for x in question.correct_answer.split('♡') if x.strip()]
                wrong_answers = [x.strip() for x in question.wrong_answer.split('♡') if x.strip()]
                
                if question.ID_QuestionType == 2:  # Текстовый ответ
                    if correct_answers and correct_answers[0] == str(user_answers):
                        correct_count += 1
                else:  # Множественный выбор
                    correct_answers_set = set(correct_answers)
                    user_answers_set = set(user_answers)
                    if (correct_answers_set.issubset(user_answers_set) and 
                        not correct_answers_set.symmetric_difference(user_answers_set).intersection(set(wrong_answers))):
                        correct_count += 1

            percentage = (correct_count / total_questions) * 100 if total_questions else 0
            grade = 'A' if percentage >= 50 else 'F'
            date_taken = datetime.now()

            # Сохраняем результат теста
            test = Tests.objects.get(ID_Test=test_id)
            user_id = request.session.get('user_id')
            
            UserTests.objects.create(
                ID_User_id=user_id,
                ID_Course_id=course_id,
                ID_Topic_id=topic_id,
                ID_Test_id=test_id,
                TestName=test.TestName,
                Result=percentage,
                Grade=grade,
                DateTaken=date_taken
            )

            return render(request, 'student/courses/test_results.html', {
                'result': {
                    'percentage': percentage,
                    'grade': grade,
                    'correct_count': correct_count,
                    'total_questions': total_questions,
                    'TestName': test.TestName
                },
                'success_message': "Результаты успешно сохранены!"
            })

        except Exception as e:
            return render(request, 'student/courses/test_results.html', 
                        {'error_message': f"Ошибка: {str(e)}"})

    return render(request, 'student/courses/test_results.html')

def user_test_results(request):
    if not request.session.get('user_id'):
        return redirect('/login/')

    user_id = request.session['user_id']
    results = UserTests.objects.filter(ID_User=user_id).select_related(
        'ID_Test', 'ID_Topic', 'ID_Course'
    ).order_by('-DateTaken')

    formatted_results = []
    for result in results:
        formatted_results.append({
            'course_name': result.ID_Course.CourseName,
            'topic_name': result.ID_Topic.TopicName,
            'test_name': result.TestName,
            'result': f"{result.Result}%",
            'grade': result.Grade,
        })

    return render(request, 'student/courses/user_test_results.html', 
                {'results': formatted_results})